#imports module. Has already been downloaded to local machine using pip install via command prompt
import openpyxl

#Creates a variable for the selected workbook.
wb = openpyxl.load_workbook(filename=r"C:\Users\Rosie\Documents\Blue_Dirt_Maps\nonblog_pages\timeline_buildings\docs\test.xlsx", use_iterators=True)
ws = wb.get_sheet_by_name('allData')

#goes through every cell in the work sheet, prints that cell + the cell in the first row  + cell in the first column (both for that particular cell)
x =2
for index, row in enumerate(ws.iter_rows('B2:DE339')):
	for cell in row:
		if cell.value:
			print(str(cell.value) + "; " 
			+str(ws.cell(row=index+2, column=1).value) + "; " 
			+str(ws.cell(row=1, column=x).value))
		if x<109:
			x+=1
		else:
			x=2